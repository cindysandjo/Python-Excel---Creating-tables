#The openpyxl module will interact with excel
import openpyxl

#The os module will navigate through the directories in our OS. Change directory to the location of your excel file
import os
os.chdir('c:\\Users\\cxsandjo\\OneDrive - Ben E Keith Company\\Desktop')

#Load the excel workbook
wb = openpyxl.load_workbook('Test.xlsx')
#define the worksheet
ws = wb.active

#View the number of rows & columns there are in the workbook
max_row = str(wb.worksheets[0].max_row)
max_col_nbr = wb.worksheets[0].max_column
max_col_ltr = openpyxl.utils.get_column_letter(max_col_nbr)
min_row = str(wb.worksheets[0].min_row)
min_col_nbr = wb.worksheets[0].min_column
min_col_ltr = openpyxl.utils.get_column_letter(min_col_nbr)

data_range = min_col_ltr+min_row+':'+max_col_ltr+max_row

print(data_range)

#Define a table style
mediumStyle = openpyxl.worksheet.table.TableStyleInfo(name='TableStyleMedium2'
                                                      ,showRowStripes=True)

#Create a table
table = openpyxl.worksheet.table.Table(ref=data_range,
                                       displayName='Customers',
                                       tableStyleInfo=mediumStyle)
#add the table to the worksheet
ws.add_table(table)

#save the workbook file
wb.save('Test.xlsx')
